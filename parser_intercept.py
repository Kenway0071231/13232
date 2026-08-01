import time
import json
import pandas as pd
from playwright.sync_api import sync_playwright

# ---------- НАСТРОЙКИ ----------
REGION = "Ярославль"
# Для теста возьмём короткий список. Потом можно расширить.
SEARCH_QUERIES = [
    "аптека", "магазин", "кафе", "ремонт", "салон красоты",
    "автосервис", "стройматериалы", "мебель", "одежда", "обувь",
    "банк", "страхование", "туризм", "гостиница", "стоматология",
    "ветеринар", "шиномонтаж", "цветы", "подарки", "зоомагазин",
    "спорт", "фитнес", "бассейн", "образование", "курсы",
    "медицинский центр", "оптика", "косметика", "парфюмерия",
    "электроника", "бытовая техника", "канцтовары", "книги",
    "хозтовары", "сантехника", "инструменты", "автозапчасти",
    "мототехника", "велосипеды", "комиссионный", "ломбард",
    "ювелирный", "часы", "ремонт обуви", "химчистка", "прачечная"
]

OUTPUT_CSV = "2gis_intercept_raw.csv"
EXCEL_FILE = "Сети_Ярославль.xlsx"

STOP_WORDS = [
    "продуктовый", "супермаркет", "гипермаркет", "продукты питания",
    "администрация", "правительство", "полиция", "суд", "прокуратура",
    "военкомат", "паспортный стол", "госуслуги", "мфц", "налоговая",
    "фсб", "мвд", "пожарная", "следственный", "избирательная",
    "законодательное", "контрольно-счетная", "городская дума", "губернатор"
]

# ---------- ФУНКЦИИ ----------
def intercept_response(response, collected):
    """Обработчик ответов: ищем данные организаций."""
    url = response.url
    # Типичные шаблоны URL внутреннего API 2GIS
    if "catalog.api.2gis.com" in url and ("search" in url or "branch" in url):
        try:
            data = response.json()
            # проверяем наличие результата
            if "result" in data and "items" in data["result"]:
                items = data["result"]["items"]
                # Преобразуем в наш формат
                for item in items:
                    point = item.get("point", {})
                    phones = []
                    for cg in item.get("contact_groups", []):
                        for c in cg.get("contacts", []):
                            if c.get("type") == "phone":
                                phones.append(c.get("value", ""))
                    phone_str = ", ".join(phones)
                    site_str = ""
                    for link in item.get("links", []):
                        if link.get("type") == "site":
                            site_str = link.get("url", "")
                            break
                    # Рубрики
                    rubrics = [r.get("name", "") for r in item.get("rubrics", [])]
                    collected.append({
                        "name": item.get("name", ""),
                        "address": item.get("address_name", ""),
                        "lat": point.get("lat"),
                        "lon": point.get("lon"),
                        "phone": phone_str,
                        "site": site_str,
                        "rubrics": rubrics,
                        "rubric_main": rubrics[0] if rubrics else ""
                    })
        except:
            pass  # не JSON или другой формат

def main():
    all_data = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # headless=False чтобы видеть процесс
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()

        # Регистрируем обработчик
        page.on("response", lambda response: intercept_response(response, all_data))

        for q in SEARCH_QUERIES:
            print(f"Ищем: {q}")
            url = f"https://2gis.ru/{REGION}/search/{q}"
            page.goto(url, wait_until="networkidle")
            time.sleep(2)  # дополнительное ожидание подгрузки

        browser.close()

    if not all_data:
        print("Не удалось перехватить данные. Попробуйте запустить с headless=False и проверьте консоль.")
        return

    print(f"Всего перехвачено записей: {len(all_data)}")

    # Убираем дубликаты по названию и адресу
    df = pd.DataFrame(all_data)
    df.drop_duplicates(subset=["name", "address"], inplace=True)
    print(f"Уникальных: {len(df)}")

    # Фильтрация по стоп-словам в рубриках
    def has_stop_word(rubrics):
        rubrics_lower = [r.lower() for r in rubrics]
        return any(any(sw in r for sw in STOP_WORDS) for r in rubrics_lower)

    mask = df["rubrics"].apply(has_stop_word)
    df = df[~mask]
    print(f"После фильтрации: {len(df)}")

    # Сохраняем CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8")
    print(f"Данные сохранены в {OUTPUT_CSV}")

    # Анализ сетей (3+ точки)
    df_unique = df.drop_duplicates(subset=["name", "address"])
    network_counts = df_unique.groupby("name")["address"].nunique()
    valid_networks = network_counts[network_counts >= 3].index
    df_networks = df_unique[df_unique["name"].isin(valid_networks)]

    # Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Сети"
    cols = ["name", "address", "phone", "site", "lat", "lon", "rubric_main", "total_points"]
    points_info = network_counts.reset_index()
    points_info.columns = ["name", "total_points"]
    df_networks = df_networks.merge(points_info, on="name", how="left")
    for col_idx, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_networks.itertuples(index=False), 2):
        for col_idx, value in enumerate(row[:len(cols)], 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(EXCEL_FILE)
    print(f"✅ Excel-файл сохранён: {EXCEL_FILE}")

if __name__ == "__main__":
    main()
