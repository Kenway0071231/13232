import time
import re
import pandas as pd
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup

# ---------- НАСТРОЙКИ ----------
REGION = "Ярославль"
SEARCH_QUERIES = [
    "аптека", "магазин", "кафе", "ремонт", "салон красоты",
    "автосервис", "стройматериалы", "мебель", "одежда", "обувь",
    "банк", "страхование", "туризм", "гостиница", "стоматология",
    "ветеринар", "шиномонтаж", "цветы", "подарки", "зоомагазин",
    "спорт", "фитнес", "бассейн", "образование", "курсы",
    "медицинский центр", "оптика", "косметика", "парфюмерия",
    "электроника", "бытовая техника", "канцтовары", "книги",
    "хозтовары", "сантехника", "инструменты", "автозапчасти",
    "мототехника", "велосипеды", "комиссионный", "ломбард",
    "ювелирный", "часы", "ремонт обуви", "химчистка", "прачечная"
]

STOP_WORDS = [
    "продуктовый", "супермаркет", "гипермаркет", "продукты питания",
    "администрация", "правительство", "полиция", "суд", "прокуратура",
    "военкомат", "паспортный стол", "госуслуги", "мфц", "налоговая",
    "фсб", "мвд", "пожарная", "следственный", "избирательная",
    "законодательное", "контрольно-счетная", "городская дума", "губернатор"
]

OUTPUT_CSV = "2gis_playwright_raw.csv"
EXCEL_FILE = "Сети_Ярославль.xlsx"

# ---------- ФУНКЦИИ ----------
def parse_page_content(html):
    """Извлекает организации из HTML страницы поиска."""
    soup = BeautifulSoup(html, 'html.parser')
    items = []

    # Селекторы для карточек организаций (могут меняться, актуальны на 2024-2025)
    # Варианты: article, div с data-testid="branch-item", div._1kf6vk
    cards = soup.find_all('article', class_=re.compile(r'_1kf6vk|_1rehek'))
    if not cards:
        cards = soup.find_all('div', attrs={'data-testid': 'branch-item'})
    if not cards:
        cards = soup.find_all('div', class_=re.compile(r'_1kf6vk'))

    for card in cards:
        name_tag = card.find('a', class_=re.compile(r'_1rehek|_1xqhjo')) or card.find('span', class_=re.compile(r'_1rehek'))
        name = name_tag.get_text(strip=True) if name_tag else ''

        address_tag = card.find('span', class_=re.compile(r'_1xqhjo|_1w9o2t'))
        address = address_tag.get_text(strip=True) if address_tag else ''

        # Рубрики (обычно span с классом _1w9o2t или подобным)
        rubric_tags = card.find_all('span', class_=re.compile(r'_1w9o2t|_1rehek'))
        rubrics = [r.get_text(strip=True) for r in rubric_tags if r.get_text(strip=True) and r != name_tag]

        # Телефон (может быть скрыт за иконкой)
        phone_tag = card.find('a', href=re.compile(r'tel:'))
        phone = phone_tag.get_text(strip=True) if phone_tag else ''

        # Сайт
        site_tag = card.find('a', href=re.compile(r'^https?://'))
        site = site_tag['href'] if site_tag else ''

        # Координаты (иногда data-атрибуты)
        lat = card.get('data-lat')
        lon = card.get('data-lon')

        items.append({
            'name': name,
            'address': address,
            'phone': phone,
            'site': site,
            'rubrics': rubrics,
            'lat': lat,
            'lon': lon
        })
    return items

def search_organizations(page, query, region):
    items = []
    url = f"https://2gis.ru/{region}/search/{query}"
    print(f"  Открываю {url}")
    page.goto(url, wait_until='networkidle')
    time.sleep(2)  # доп. ожидание загрузки

    # Попытка закрыть всплывающие окна (если есть)
    try:
        page.click('button[aria-label="Закрыть"]', timeout=1000)
    except:
        pass

    page_num = 1
    while True:
        print(f"    Страница {page_num}")
        html = page.content()
        new_items = parse_page_content(html)
        if not new_items:
            print("    Ничего не найдено на этой странице.")
            break
        items.extend(new_items)

        # Пагинация: ищем кнопку следующей страницы
        next_btn = page.query_selector('a[data-page]')
        if not next_btn:
            next_btn = page.query_selector('button[aria-label="Следующая страница"]')
        if not next_btn:
            next_btn = page.query_selector('text=Следующая')
        if not next_btn:
            # Пробуем найти любой элемент с data-page на 1 больше
            next_btn = page.query_selector(f'a[data-page="{page_num+1}"]')

        if next_btn:
            next_btn.click()
            page_num += 1
            time.sleep(1.5)
            page.wait_for_load_state('networkidle')
        else:
            print("    Пагинация не найдена, завершаю.")
            break
    return items

def main():
    all_data = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # headless=False для отладки
        context = browser.new_context(
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        page = context.new_page()

        for query in SEARCH_QUERIES:
            print(f"\nОбрабатываю запрос: {query}")
            try:
                results = search_organizations(page, query, REGION)
                print(f"  Найдено организаций: {len(results)}")
                all_data.extend(results)
            except Exception as e:
                print(f"  Ошибка: {e}")
            time.sleep(2)  # пауза между запросами

        browser.close()

    if not all_data:
        print("Нет данных для сохранения.")
        return

    # Убираем дубликаты
    df = pd.DataFrame(all_data)
    df.drop_duplicates(subset=['name', 'address'], inplace=True)
    print(f"Всего уникальных организаций: {len(df)}")

    # Фильтрация по стоп-словам в рубриках
    def has_stop_word(rubrics):
        rubrics_lower = [r.lower() for r in rubrics]
        return any(any(sw in r for sw in STOP_WORDS) for r in rubrics_lower)

    mask = df['rubrics'].apply(has_stop_word)
    df = df[~mask]
    print(f"После фильтрации стоп-слов: {len(df)}")

    # Сохраняем CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding='utf-8')
    print(f"Данные сохранены в {OUTPUT_CSV}")

    # Анализ сетей (3+ точки)
    df['rubric_main'] = df['rubrics'].apply(lambda x: x[0] if x else '')
    df_unique = df.drop_duplicates(subset=['name', 'address'])
    network_counts = df_unique.groupby('name')['address'].nunique()
    valid_networks = network_counts[network_counts >= 3].index
    df_networks = df_unique[df_unique['name'].isin(valid_networks)]

    # Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Сети"
    cols = ['name', 'address', 'phone', 'site', 'lat', 'lon', 'rubric_main', 'total_points']
    points_info = network_counts.reset_index()
    points_info.columns = ['name', 'total_points']
    df_networks = df_networks.merge(points_info, on='name', how='left')
    for col_idx, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_networks.itertuples(index=False), 2):
        for col_idx, value in enumerate(row[:len(cols)], 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(EXCEL_FILE)
    print(f"✅ Excel-файл сохранён: {EXCEL_FILE}")

if __name__ == '__main__':
    main()
