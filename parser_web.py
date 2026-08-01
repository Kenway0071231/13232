import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import re
import json

# ---------- НАСТРОЙКИ ----------
REGION_NAME = "Ярославль"   # город, по которому ищем
# Широкий список поисковых запросов (можно добавить свои)
SEARCH_QUERIES = [
    "аптека", "магазин", "кафе", "ремонт", "салон красоты",
    "автосервис", "стройматериалы", "мебель", "одежда", "обувь",
    "банк", "страхование", "туризм", "гостиница", "стоматология",
    "ветеринар", "шиномонтаж", "цветы", "подарки", "зоомагазин",
    "спорт", "фитнес", "бассейн", "образование", "курсы",
    "медицинский центр", "оптика", "косметика", "парфюмерия",
    "электроника", "бытовая техника", "канцтовары", "книги",
    "хозтовары", "сантехника", "инструменты", "автозапчасти",
    "мототехника", "велосипеды", "комиссионный", "ломбард",
    "ювелирный", "часы", "ремонт обуви", "химчистка", "прачечная"
]

OUTPUT_CSV = "2gis_web_raw.csv"
EXCEL_FILE = "Сети_Ярославль.xlsx"

STOP_WORDS = [
    "продуктовый", "супермаркет", "гипермаркет", "продукты питания",
    "администрация", "правительство", "полиция", "суд", "прокуратура",
    "военкомат", "паспортный стол", "госуслуги", "мфц", "налоговая",
    "фсб", "мвд", "пожарная", "следственный", "избирательная",
    "законодательное", "контрольно-счетная", "городская дума", "губернатор"
]

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
}

# ---------- ФУНКЦИИ ----------
def fetch_page(url, params=None):
    for attempt in range(3):
        try:
            resp = requests.get(url, headers=HEADERS, params=params, timeout=30)
            if resp.status_code == 200:
                return resp
            else:
                print(f"    Ошибка {resp.status_code} для {url}")
                time.sleep(2)
        except Exception as e:
            print(f"    Ошибка: {e}")
            time.sleep(5)
    return None

def parse_branch_card(card):
    """Извлечение данных из карточки организации на странице поиска."""
    name_tag = card.find('a', class_='_1rehek')
    if not name_tag:
        name_tag = card.find('span', class_='_1rehek')
    name = name_tag.get_text(strip=True) if name_tag else ''

    address_tag = card.find('span', class_='_1xqhjo')
    address = address_tag.get_text(strip=True) if address_tag else ''

    # Рубрики
    rubric_tags = card.find_all('span', class_='_1w9o2t')
    rubrics = [r.get_text(strip=True) for r in rubric_tags]

    # Телефон
    phone_tag = card.find('a', class_='_2lcm398')
    phone = phone_tag.get_text(strip=True) if phone_tag else ''

    # Сайт
    site_tag = card.find('a', class_='_1rehek', href=re.compile(r'^https?://'))
    site = site_tag['href'] if site_tag else ''

    # Координаты (иногда в data-атрибутах)
    lat, lon = None, None
    coord_tag = card.find(attrs={"data-lat": True})
    if coord_tag:
        lat = coord_tag.get('data-lat')
        lon = coord_tag.get('data-lon')

    return {
        'name': name,
        'address': address,
        'phone': phone,
        'site': site,
        'rubrics': rubrics,
        'lat': lat,
        'lon': lon
    }

def search_organizations(query, region):
    items = []
    page = 1
    while True:
        print(f"  Поиск '{query}', страница {page}...")
        url = f"https://2gis.ru/{region}/search/{query}"
        params = {'page': page} if page > 1 else None
        resp = fetch_page(url, params)
        if not resp:
            break

        soup = BeautifulSoup(resp.text, 'html.parser')
        cards = soup.find_all('div', class_='_1kf6vk')
        if not cards:
            cards = soup.find_all('article', class_='_1kf6vk')  # возможно, другой класс
        if not cards:
            # попробуем найти по data-testid
            cards = soup.find_all('div', attrs={'data-testid': 'branch-item'})

        if not cards:
            print("    Организации не найдены (возможно, изменилась вёрстка).")
            break

        for card in cards:
            data = parse_branch_card(card)
            items.append(data)

        # Проверяем, есть ли следующая страница
        next_btn = soup.find('a', class_='_1rehek', attrs={'data-page': str(page+1)})
        if not next_btn:
            # альтернативный поиск пагинации
            pagination = soup.find('div', class_='_1kf6vk')
            if pagination:
                next_link = pagination.find('a', string=str(page+1))
                if not next_link:
                    break
            else:
                break
        page += 1
        time.sleep(1.5)   # пауза между страницами

    return items

def main():
    all_data = []
    for query in SEARCH_QUERIES:
        print(f"\nОбрабатываю запрос: {query}")
        results = search_organizations(query, REGION_NAME)
        print(f"  Найдено организаций: {len(results)}")
        all_data.extend(results)
        time.sleep(2)  # между запросами

    # Убираем дубликаты (по названию и адресу)
    df = pd.DataFrame(all_data)
    df.drop_duplicates(subset=['name', 'address'], inplace=True)
    print(f"Всего уникальных организаций: {len(df)}")

    # Фильтрация по стоп-словам в рубриках
    def has_stop_word(rubrics):
        rubrics_lower = [r.lower() for r in rubrics]
        return any(any(sw in r for sw in STOP_WORDS) for r in rubrics_lower)

    mask = df['rubrics'].apply(has_stop_word)
    df = df[~mask]
    print(f"После фильтрации стоп-слов: {len(df)}")

    # Сохраняем CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding='utf-8')
    print(f"Данные сохранены в {OUTPUT_CSV}")

    # Анализ сетей (3+ точки)
    df['rubric_main'] = df['rubrics'].apply(lambda x: x[0] if x else '')
    df_unique = df.drop_duplicates(subset=['name', 'address'])
    network_counts = df_unique.groupby('name')['address'].nunique()
    valid_networks = network_counts[network_counts >= 3].index
    df_networks = df_unique[df_unique['name'].isin(valid_networks)]

    # Excel
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Сети"
    cols = ['name', 'address', 'phone', 'site', 'lat', 'lon', 'rubric_main', 'total_points']
    # Добавим total_points
    points_info = network_counts.reset_index()
    points_info.columns = ['name', 'total_points']
    df_networks = df_networks.merge(points_info, on='name', how='left')
    for col_idx, col_name in enumerate(cols, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_networks.itertuples(index=False), 2):
        for col_idx, value in enumerate(row[:len(cols)], 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(EXCEL_FILE)
    print(f"✅ Excel-файл сохранён: {EXCEL_FILE}")

if __name__ == '__main__':
    main()
